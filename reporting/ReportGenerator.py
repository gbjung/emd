import time
import string
import copy
from bs4 import BeautifulSoup
from operator import itemgetter
from django.conf import settings

from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from salesforce.SalesforceClient import SalesforceClient
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment


class ReportGenerator:

    def __init__(self):
        sfc = SalesforceClient()
        self.sf = sfc.sf
        self.use_activities = True
        self.incumbent_map = {}
        double_border_side = Side(border_style="thin")
        self.square_border = Border(left=double_border_side,
                                right=double_border_side,
                                top=double_border_side,
                                bottom=double_border_side)

    def fetch_events(self, account_id):
        '''
        Not used anymore, keep in case they want it again
        '''
        sf_fields = "Subject, ActivityDate"
        today = date.today()
        where = "(AccountId ='{}' and ActivityDate >= {})".format(account_id, str(today))
        order = " order by ActivityDate asc"
        query = "SELECT "+ sf_fields + " From Event WHERE " + where + order
        return self.sf.query(query)

    def add_account_info(self, sheet, account_info):
        sheet['B1'] = account_info['Name']
        sheet['B2'] = account_info['Pipeline_Date__c'][:10] if account_info['Pipeline_Date__c'] else ''
        sheet['B3'] = account_info['Following_Pipeline_Date__c'][:10] if account_info['Following_Pipeline_Date__c'] else ''
        sheet['H1'] = account_info['Target_Customers__c']
        sheet['H2'] = account_info['Target_Value__c']
        sheet['H3'] = account_info['Other_Targets__c']

    def add_event_info(self, sheet, events):
        '''
        Not used anymore, keep in case they want it again
        '''
        if events['totalSize']:
            event = events['records'][0]
            sheet['B2'] = '{} ({})'.format(event['Subject'], event['ActivityDate'])

        if events['totalSize'] > 1:
            event = events['records'][1]
            sheet['B3'] = '{} ({})'.format(event['Subject'], event['ActivityDate'])

    def fetch_opportunities(self, account_id):
        sf_fields = ', '.join(['Name', 'StageName', 'Competition_Type__c', 'Contract_Vehicle__c',
                     'Incumbent_BGOV_Page__c', 'Contract__c', 'Amount',
                     'Client_Role__c', 'Incumbent__c', 'CloseDate', 'Executive_Summary__c',
                     'Reasons_to_Pursue__c', 'Challenges__c', 'FBO_Page__c',
                     'Place_of_Performance__c', 'NAICS_Code__c', 'Contract_End_Date__c',
                     'Eligible_to_Bid__c', 'Projected_Award_Date__c', 'Requirements__c',
                     'Solicitation_Number__c'])
        if self.use_activities:
            extra_type_fields = (', (select Subject, Description, Status, LastModifiedDate from Tasks)')
        else:
            extra_type_fields = (', (select Name, Update__c, Update_Type__c, CreatedDate from Opportunity_Updates__r)')
        teaming_partners_query = ', (select Name, Justification__c, Status__c, Point_of_Contact__c from Potential_Teaming_Partners__r)'
        extra_type_fields += teaming_partners_query
        where = "(StageName not in ('Monitoring', 'SP 7 - RFP Submitted', 'No Go') and AccountId ='{}')".format(account_id)
        order = " order by CloseDate asc"
        query = "SELECT " + sf_fields + extra_type_fields + " FROM Opportunity WHERE " + where + order
        return self.sf.query(query)

    def add_opportunities(self, sheet, opportunities):
        row = 5
        counter = 1
        for op in opportunities:
            sheet['A{}'.format(row)] = counter
            sheet['B{}'.format(row)] = op['Name']
            sheet['C{}'.format(row)] = op['StageName']
            sheet['D{}'.format(row)] = op['Solicitation_Number__c']
            sheet['E{}'.format(row)] = op['Competition_Type__c']
            sheet['F{}'.format(row)] = op['Contract_Vehicle__c']
            sheet['G{}'.format(row)] = op['Contract__c']
            sheet['H{}'.format(row)] = op['Amount']
            sheet['H{}'.format(row)].number_format = '"$"#,##0.00'
            sheet['I{}'.format(row)] = op['Client_Role__c']
            sheet['J{}'.format(row)] = self.find_incumbent(op['Incumbent__c'])
            sheet['L{}'.format(row)] = op['CloseDate']
            for col in string.ascii_uppercase[:12]:
                sheet['{}{}'.format(col,row)].border = self.square_border
                sheet['{}{}'.format(col,row)].alignment = Alignment(wrap_text=True,
                                                                    vertical='center',
                                                                    horizontal='center')
            row += 1
            counter += 1

    def find_incumbent(self, incumbent_id):
        if not incumbent_id:
            return ''

        if incumbent_id in self.incumbent_map:
            return self.incumbent_map.get(incumbent_id)

        account = self.sf.Account.get(incumbent_id)
        self.incumbent_map[incumbent_id] = account['Name']
        return account['Name']

    def format_to_excel(self, workbook, account_info, opportunities):
        sheet = workbook["Current Pipeline"]
        self.add_account_info(sheet, account_info)
        self.add_opportunities(sheet, opportunities)
        account_updates = self.fetch_account_updates(account_info)
        self.add_account_updates(workbook, account_updates)

        return workbook

    def fetch_account_updates(self, account_info):
        two_weeks_later = (date.today() - relativedelta(weeks=2, days=1)).strftime('%Y-%m-%dT%H:%M:%SZ')
        query = "SELECT Subject, Description, Status, Priority, ActivityDate FROM Task WHERE (WhatId='{}' and CreatedDate >= {})".format(account_info['Id'], two_weeks_later)
        return self.sf.query(query)


    def add_account_updates(self, workbook, account_updates):
        sheet = workbook["Account Updates"]
        row = 2
        for update in account_updates['records']:
            sheet['A{}'.format(row)] = update['Subject']
            sheet['B{}'.format(row)] = update['Description']
            sheet['C{}'.format(row)] = update['Status']
            sheet['D{}'.format(row)] = update['Priority']
            sheet['E{}'.format(row)] = update['ActivityDate']
            for col in ['A','B','C','D','E']:
                sheet['{}{}'.format(col,row)].border = self.square_border
                sheet['{}{}'.format(col,row)].alignment = Alignment(wrap_text=True,
                                                                    vertical='center',
                                                                    horizontal='center')
            row += 1

    def add_ops(self, workbook, opportunities, file_name, account_name):
        row = 1
        for opp in opportunities:
            original_sheet = workbook['Opp']
            sheet = workbook.copy_worksheet(original_sheet)
            sheet_name = 'Opp {}'.format(row)
            sheet.title = sheet_name
            self.format_op(sheet, opp, account_name)
            row += 1
            workbook.save(filename=file_name)

        workbook.remove_sheet(workbook['Opp'])
        workbook.save(filename=file_name)
        return workbook

    def seperate_update_types(self, updates):
        stores = {}

        for update in updates['records']:
            update['CreatedDate'] = datetime.strptime(update['CreatedDate'][:10],
                                                      '%Y-%m-%d').strftime('%Y-%m-%d')
            update_c = update['Update__c'] or ''
            update['Update__c'] = BeautifulSoup(update_c, "html.parser").text
            if update['CreatedDate'] in stores:
                stores[update['CreatedDate']].append(update)
            else:
                stores[update['CreatedDate']] = [update]

        for store in stores:
            stores[store] = sorted(stores[store], key=itemgetter('CreatedDate'), reverse=True)

        return stores

    def format_op_updates(self, sheet, opp):
        # V1 used op updates, keep around incase they wanna use it l8r
        row = 18
        updates = opp['Opportunity_Updates__r']

        if not updates:
            return row

        stores = self.seperate_update_types(updates)
        for date, updates in sorted(stores.items()):
            sheet.merge_cells('A{}:D{}'.format(row, row))
            sheet['A{}'.format(row)] = date
            sheet['A{}'.format(row)].alignment = Alignment(horizontal='center')
            sheet['A{}'.format(row)].font = Font(bold=True)
            for col in ['A','B','C','D']:
                sheet['{}{}'.format(col,row)].border = self.square_border
            row += 1
            for update in updates:
                sheet['A{}'.format(row)] = update['Update_Type__c']
                sheet['A{}'.format(row)].border = self.square_border
                sheet.merge_cells('B{}:D{}'.format(row, row))
                sheet['B{}'.format(row)] = update['Update__c']
                sheet['B{}'.format(row)].font = Font(bold=False)
                for col in ['B','C','D']:
                    sheet['{}{}'.format(col,row)].border = self.square_border
                row += 1
        return row

    def format_op(self, sheet, opp, account_name):
        sheet['B2'] = opp['Name'] #Opportunity Title
        sheet['D2'] = account_name #Customer
        sheet['B4'] = opp['Executive_Summary__c'] #Executive Summary
        sheet['B5'] = opp['Reasons_to_Pursue__c'] #Reason to Pursue
        sheet['B6'] = opp['Challenges__c'] # Challenges
        sheet['B8'] = opp['Client_Role__c'] # Client Role
        sheet['B9'] = opp['CloseDate'] # RFP Date
        sheet['B10'] = opp['Projected_Award_Date__c'] # Award Date
        sheet['B11'] = opp['Amount'] # Value
        sheet['B11'].number_format = '"$"#,##0.00'
        sheet['B11'].alignment = Alignment(horizontal='left')
        sheet['B13'] = opp['Contract__c'] # Contract length
        sheet['B14'] = opp['Competition_Type__c'] # Competition Type
        if opp['Requirements__c']:
            sheet['B16'] = BeautifulSoup(opp['Requirements__c'], "html.parser").text # Main Requirements
        sheet['D3'] = opp['FBO_Page__c'] # FBO.gov Link
        sheet['D8'] = opp['Contract_Vehicle__c'] # Contract Vehicle
        sheet['D9'] = opp['Place_of_Performance__c'] # Place of Performance
        sheet['D10'] = opp['NAICS_Code__c'] # NAICS Code
        sheet['D12'] = self.find_incumbent(opp['Incumbent__c']) # Incumbent
        sheet['D13'] = opp['Contract_End_Date__c'] # Contract End Date
        sheet['D14'] = opp['Eligible_to_Bid__c'] # Eligible to Bid
        if self.use_activities:
            row = self.format_op_activites(sheet, opp)
        else:
            row = self.format_op_updates(sheet, opp)
        self.format_teaming_partners(sheet, opp, row)

    def format_teaming_partners(self, sheet, opp, row):
        row = self.format_teaming_partners_headers(sheet, opp, row)
        partners = opp['Potential_Teaming_Partners__r']

        if not partners:
            return

        for partner in partners['records']:
            sheet['A{}'.format(row)] = partner['Name']
            if partner['Justification__c']:
                sheet['B{}'.format(row)] = BeautifulSoup(partner['Justification__c'], "html.parser").text.strip()
                sheet['B{}'.format(row)].border = self.square_border
                sheet['B{}'.format(row)].alignment = Alignment(wrap_text=True,
                                                               vertical='center',
                                                               horizontal='left')
            sheet['C{}'.format(row)] = partner['Status__c']
            sheet['D{}'.format(row)] = self.get_POC(partner['Point_of_Contact__c'])
            for col in ['A', 'C','D']:
                sheet['{}{}'.format(col,row)].border = self.square_border
                sheet['{}{}'.format(col,row)].alignment = Alignment(wrap_text=True,
                                                                    vertical='center',
                                                                    horizontal='center')
            row += 1

    def get_POC(self, poc_id):
        if not poc_id:
            return None
        account_info = self.sf.Contact.get(poc_id)
        return account_info['Name']

    def format_teaming_partners_headers(self, sheet, opp, row):
        sheet.merge_cells('A{}:D{}'.format(row, row))
        sheet['A{}'.format(row)] = "Potential Teaming Partners"
        ref = sheet['A17']._style
        sheet['A{}'.format(row)]._style = ref
        for col in ['A','B','C','D']:
            sheet['{}{}'.format(col,row)].border = self.square_border
        row += 1
        headers = [('A', 'Potential Teaming Partner'),
                   ('B', 'Justification'),
                   ('C', 'Status'),
                   ('D', 'Point of Contact')]
        ref = sheet['A2']._style
        for col, name in headers:
            sheet['{}{}'.format(col, row)] = name
            sheet['{}{}'.format(col, row)]._style = ref
            sheet['{}{}'.format(col, row)].alignment = Alignment(horizontal='center')
        row += 1
        return row

    def format_op_activites(self, sheet, opp):
        self.format_op_activites_headers(sheet, opp)
        row = 19
        activities = opp['Tasks']

        if not activities:
            return row

        for activity in activities['records']:
            sheet['A{}'.format(row)] = activity['Subject']
            sheet['B{}'.format(row)] = activity['Description']
            sheet['B{}'.format(row)].alignment = Alignment(wrap_text=True,
                                                           vertical='center',
                                                           horizontal='left')
            sheet['C{}'.format(row)] = activity['Status']
            date = datetime.strptime(activity['LastModifiedDate'][:10],
                                     '%Y-%m-%d').strftime('%-m/%-d/%Y')
            sheet['D{}'.format(row)] = date
            for col in ['A','C','D']:
                sheet['{}{}'.format(col,row)].border = self.square_border
                sheet['{}{}'.format(col,row)].alignment = Alignment(wrap_text=True,
                                                                    vertical='center',
                                                                    horizontal='center')
            row += 1
        return row

    def format_op_activites_headers(self, sheet, opp):
        row = 18
        headers = [('A', 'Update'),
                   ('B', 'Details'),
                   ('C', 'Status'),
                   ('D', 'Due Date')]
        ref = sheet['A2']._style
        for col, name in headers:
            sheet['{}{}'.format(col, row)] = name
            sheet['{}{}'.format(col, row)]._style = ref
            sheet['{}{}'.format(col, row)].alignment = Alignment(horizontal='center')

    def generate_report(self, account_id):
        account_info = self.sf.Account.get(account_id)
        account_name = account_info['Name']
        file_name = account_name.replace(" ", "") + 'Report.xlsx'
        opportunities = self.fetch_opportunities(account_id)['records']
        workbook = load_workbook(settings.BASE_DIR + '/pipeline_updates.xlsx')
        workbook = self.format_to_excel(workbook, account_info, opportunities)
        workbook = self.add_ops(workbook, opportunities, file_name, account_name)
        return file_name
